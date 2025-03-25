import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

FILE_PATH = "data.xlsx"

def load_excel_data():
    """Load data from the Excel file and clean it."""
    try:
        data = pd.read_excel(FILE_PATH, sheet_name=None)
        cleaned_data = {}
        for sheet, df in data.items():
            df_cleaned = df.dropna(how='all').reset_index(drop=True)
            df_cleaned = df_cleaned.loc[:, ~df_cleaned.columns.str.contains('^Unnamed')]
            cleaned_data[sheet] = df_cleaned
        return cleaned_data
    except FileNotFoundError:
        print("Файл не найден.")
        return None


def get_students():
    data = load_excel_data()
    if "Успеваемость студентов" not in data:
        return []

    students_df = data["Успеваемость студентов"].dropna(how="all").reset_index(drop=True)

    # Автоматически определим, где начинаются реальные данные
    for index, row in students_df.iterrows():
        if "ФИО студента" in row.values:
            students_df.columns = row  # Устанавливаем правильные заголовки
            students_df = students_df.iloc[index + 1:].reset_index(drop=True)
            break

    # Преобразуем данные в словарь, заменяя NaN на пустые строки
    students_dict = students_df.to_dict(orient="records")
    students_dict = [{k: ("" if pd.isna(v) else v) for k, v in record.items()} for record in students_dict]

    return students_dict

def get_professors():
    """Return all available professor data."""
    data = load_excel_data()
    if "Информация о сотрудниках" not in data:
        print("Лист с данными о преподавателях не найден.")
        return []

    professors_df = data["Информация о сотрудниках"].dropna(how="all").reset_index(drop=True)

    for index, row in professors_df.iterrows():
        if "ФИО преподавателя/сотрудника" in row.values:
            professors_df.columns = row
            professors_df = professors_df.iloc[index + 1:].reset_index(drop=True)
            break

    return professors_df.to_dict(orient="records")


def save_data(students=None, professors=None):
    """Save students and professors data while preserving all sheets and styles."""
    try:
        # Загружаем текущую книгу Excel с сохранением форматирования
        workbook = load_workbook(FILE_PATH)

        # Очистка данных студентов перед сохранением
        if students is not None:
            if "Успеваемость студентов" not in workbook.sheetnames:
                sheet = workbook.create_sheet("Успеваемость студентов")
            else:
                # Очищаем существующие данные и заголовки
                workbook.remove(workbook["Успеваемость студентов"])
                sheet = workbook.create_sheet("Успеваемость студентов")

            students_df = pd.DataFrame(students)
            for row in dataframe_to_rows(students_df, index=False, header=True):
                sheet.append(row)

        # Очистка данных преподавателей перед сохранением
        if professors is not None:
            if "Информация о сотрудниках" not in workbook.sheetnames:
                sheet = workbook.create_sheet("Информация о сотрудниках")
            else:
                # Полное удаление листа и создание нового
                workbook.remove(workbook["Информация о сотрудниках"])
                sheet = workbook.create_sheet("Информация о сотрудниках")

            professors_df = pd.DataFrame(professors)
            for row in dataframe_to_rows(professors_df, index=False, header=True):
                sheet.append(row)

        # Сохраняем обновленный файл
        workbook.save(FILE_PATH)

    except PermissionError:
        print("Ошибка: Файл открыт в другой программе. Закройте его и повторите попытку.")
